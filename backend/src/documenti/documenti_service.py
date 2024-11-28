from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Optional
from datetime import datetime
from ..raccolte.raccolte_service import (
    find_raccolte,
    find_raccolte_aggregate,
    segna_raccolte_esporate,
    Raccolta,
)

capacita_contenitori = {
    "FP": 10,
    "TN": 1.5,
    "FP": 10,
    "TN": 1.5,
    "FF": 15,
    "SC": 1.5,
    "FF": 15,
    "SC": 1.5,
    "FC": 12,
    "IBC": 57,
    "CARTA": 1,
    "PLASTICA": 1,
    "LEGNO": 1,
}


async def report_raccolte_byte_buffer(
    session: AsyncSession,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    esportato: bool = None,
) -> BytesIO:
    raccolte: list[dict] = await find_raccolte_aggregate(
        session, start_date, end_date, esportato=esportato
    )
    # create an excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Codice", "HP", "UNIMIS", "QTAMOV", "Raggruppamento"])
    column_widths = [20, 20, 10, 10, 15]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    for raccolta in raccolte:
        (
            codice_eer,
            quantita,
            codice_raggruppamento,
            um,
            codice_pittogramma,
        ) = raccolta
        sheet.append(
            [
                codice_eer,
                codice_pittogramma,
                um,
                quantita,
                codice_raggruppamento,
            ]
        )

    # save the workbook to a bytes buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    await segna_raccolte_esporate(session, start_date, end_date, esportato=esportato)

    return buffer


async def report_custom(
    session: AsyncSession, start_date: Optional[datetime], end_date: Optional[datetime]
) -> BytesIO:
    raccolte: list[Raccolta] = await find_raccolte(
        session, start_date, end_date, eager_mode=True
    )
    # create an excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active

    sheet.title = "Resoconto scarico"
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 20
    # write the column headers
    sheet.append(["Data", "Codice EER", "N Contenitori", "Peso (kg)"])

    for raccolta in raccolte:
        sheet.append(
            [
                raccolta.data.strftime("%d/%m/%Y %H:%M"),
                raccolta.codice_eer,
                raccolta.quantita,
                raccolta.quantita * capacita_contenitori[raccolta.rifiuto.contenitore],
            ]
        )

    # save the workbook to a bytes buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
